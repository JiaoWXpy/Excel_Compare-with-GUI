'''
@Author  :   {RockyChen}

@Contact :   {rocky_chen777@163.com}

@Script:   ${Excel Compare Tool}

@File    :   ${homepage}.py

@Time    :   ${2019} ${3}${14}

---------------------
'''
热土
import wx
import openpyxl
import traceback
import wx.grid
import operator
from openpyxl.utils import get_column_letter

########################################################################
# 设置读取文件后缀类型
wildcard1 = "Excel source (*.xlsx)|*.xlsx"
wildcard2 = "Excel source (*.xlsx)|*.xlsx"
########################################################################
class MyForm(wx.Frame):

    # -------------------------------------------------------------------
    # set the window layout
    def __init__(self):
        wx.Frame.__init__(self, None, wx.ID_ANY,"Beyond Compare Pefessional 3.0",pos=(0, 0), size=(700,700))
        # 全局变量
        global TxtCfn
        # Frame布局
        self.panel = wx.Panel(self, wx.ID_ANY)
        TxtCfn = wx.TextCtrl(self.panel, pos=(15, 5), size=(200, 25))

        # 居中
        self.Center()

        self.dic = {}  # 单元格变动
        self.common_sheet = []
        self.current_sheet = []
        self.cg_col = {}  # 列增删

        # 菜单栏
        menu = wx.Menu()
        contract = menu.Append(-1, 'Main Menu', '   欢迎使用 Beyond Compare 3.0', )
        menu.AppendSeparator()
        exit = menu.Append(-1, 'Exit','   单机退出')
        self.Bind(wx.EVT_MENU, self.OnContract, contract)
        self.Bind(wx.EVT_MENU, self.OnExit, exit)
        menuBar = wx.MenuBar()
        menuBar.Append(menu, '   主菜单')
        self.SetMenuBar(menuBar)

        # 创建按钮：openA/openB/readSheet/compare
        btnCA = wx.Button(self.panel, label="File A", pos=(225, 5), size=(40, 25))
        btnCB = wx.Button(self.panel, label="File B", pos=(270, 5), size=(40, 25))
        btnCC = wx.Button(self.panel, label="Read Sheet", pos=(315, 5), size=(73, 25))
        compare_excel = wx.Button(self.panel, label="Compare Excel", pos=(490, 5), size=(120, 25))

        # 左上角图片
        icon_1 = wx.Icon(name='C:\\Users\\Administrator\\Desktop\\rocky_exe\\bp.png', type=wx.BITMAP_TYPE_PNG)
        self.SetIcon(icon_1)

        # 光标设置
        self.SetToolTip(wx.ToolTip('Beyond Compare'))
        self.SetCursor(wx.StockCursor(wx.CURSOR_BULLSEYE))

        # 下方状态栏
        self.CreateStatusBar()

        # 创建一个ComboBox对象
        self.list = ['行增删', '列增删', '单元格改动']
        cb2 = wx.ComboBox(self.panel, 1, choices=self.list, style=wx.CB_READONLY,pos = (15,35),size = (100,100))
        self.Bind(wx.EVT_COMBOBOX, self.on_cbbox, cb2)

        # 中间表格展示区
        image_file_1 = 'C:\\Users\\Administrator\\Desktop\\rocky_exe\\new.png'
        to_bmp_image_1 = wx.Image(image_file_1, wx.BITMAP_TYPE_ANY).ConvertToBitmap()
        self.bitmap_1 = wx.StaticBitmap(self.panel, -1, to_bmp_image_1, pos = (20,70),size = (260,260))

        image_file_2 = 'C:\\Users\\Administrator\\Desktop\\rocky_exe\\new.png'
        to_bmp_image_2 = wx.Image(image_file_2, wx.BITMAP_TYPE_ANY).ConvertToBitmap()
        self.bitmap_2 = wx.StaticBitmap(self.panel, -1, to_bmp_image_2, pos=(400,70), size=(260, 260))

        # 按钮事件绑定
        btnCA.Bind(wx.EVT_BUTTON, self.onOpenFileA)
        btnCB.Bind(wx.EVT_BUTTON, self.onOpenFileB)
        btnCC.Bind(wx.EVT_BUTTON, self.readSheet)
        compare_excel.Bind(wx.EVT_BUTTON, self.compareExcel)

    # -------------------------------------------------------------------
    # def [onOpenFileA] 打开A.xlsx文件
    def onOpenFileA(self, event):
        """
        创建并展示文件对话框
        :param event:单击openBfile
        """
        dlg = wx.FileDialog(self, message="Choose a file",defaultFile="",wildcard=wildcard1,style=wx.FD_OPEN | wx.FD_MULTIPLE | wx.FD_CHANGE_DIR)
        if dlg.ShowModal() == wx.ID_OK:
            tmp = ""

            paths = dlg.GetPaths()
            for path in paths:
                tmp = tmp + path

            # 输出文件路径
            TxtCfn.SetValue(tmp)

            f1 = openpyxl.load_workbook(TxtCfn.GetValue(), read_only=False,
                  data_only=False, keep_links=True)
            self.f1 = f1
        dlg.Destroy()

    # -------------------------------------------------------------------
    # def [onOpenFileA] 打开B.xlsx文件
    def onOpenFileB(self, event):
        """
        创建并展示文件对话框
        :param event:单击openAfile
        """
        dlg = wx.FileDialog(self, message="Choose a file",defaultFile="",wildcard=wildcard1,style=wx.FD_OPEN | wx.FD_MULTIPLE | wx.FD_CHANGE_DIR)
        if dlg.ShowModal() == wx.ID_OK:
            tmp = ""
            # paths = dlg.GetPaths()
            paths = dlg.GetPaths()
            # print "You chose the following file(s):"
            for path in paths:
                tmp = tmp + path
            # set the value of TextCtrl[filename]
            TxtCfn.SetValue(tmp)
            f2 = openpyxl.load_workbook(TxtCfn.GetValue(), read_only=False,
                                        data_only=False, keep_links=True)
            self.f2 = f2
        dlg.Destroy()

    # -------------------------------------------------------------------
    # def [on_cbbox] 展示比对内容
    def on_cbbox(self, event):
        """
         根据ComboBox中选择的比对内容，展示A、B文件差异
         :param event:选择比对内容的combobox
        """
        # 当前ComboBox中选项从0开始
        self.current_selection = event.GetSelection()

        # **************************************************
        # 单元格改动
        if self.current_selection == 2:

            # 清空历史数据
            self.gd.ClearGrid()

            # 打印表头
            self.gd.SetCellValue(0,0,'坐标')
            self.gd.SetCellValue(0,1,'旧值')
            self.gd.SetCellValue(0,2,'新值')

            key_list = list(self.dic.keys())

            #把列数字转换成Excel表中对应的字母
            key_word_list = []
            for key in key_list:
                key_word_list.append(((key[0]),get_column_letter(key[1])))
            print(key_word_list)
            print(key_list)

            # 循环打印字典中内容

            for row in range(1,self.dic.keys().__len__()+1):
              #  print(str((key_list[row - 1][0]), get_column_letter(key_list[row - 1][1])))
              #  print(str(key_list[row - 1]))
                #self.gd.SetCellValue(row, 0, (str(key_list[row - 1]))
                self.gd.SetCellValue(row, 0, (str(key_word_list[row - 1])))
                self.gd.SetCellValue(row, 1, (self.dic[key_list[row - 1]]).split('-')[0])
                self.gd.SetCellValue(row, 2, (self.dic[key_list[row - 1]]).split('-')[1])



        # **************************************************
        # 行增删
        elif self.current_selection == 0:
            self.gd.ClearGrid()
            self.gd.SetCellValue(0, 0, '行号')
            self.gd.SetCellValue(0, 1, '改动')

            ls = list(self.cg_row.items())

            for item in range(1,len(self.cg_row)+1):
                self.gd.SetCellValue(item,0,str(ls[item-1][0]))
                self.gd.SetCellValue(item,1,str(ls[item-1][1]))

            # 表传送功能
            self.gd.Bind(wx.grid.EVT_GRID_SELECT_CELL, self.onRowSelect)
        # **************************************************
        # 列增删
        elif self.current_selection == 1:
            self.gd.ClearGrid()

            self.gd.SetCellValue(0, 0, '列号')
            self.gd.SetCellValue(0, 1, '改动')

            ls = list(self.cg_col.items())

            for item in range(1, len(self.cg_col) + 1):
                self.gd.SetCellValue(item, 0, str(ls[item - 1][0]))
                self.gd.SetCellValue(item, 1, str(ls[item - 1][1]))

            # 表传送功能
            self.gd.Bind(wx.grid.EVT_GRID_SELECT_CELL, self.onColSelect)

    # -------------------------------------------------------------------
    # def [readSheet] 选择sheet
    def readSheet(self,event):
        """
        读取A、B文件中共有的Sheet，并选择
        :param event:单击combobox中的某个sheet
        """

        # 取A、B文件共有的sheet
        for i in self.f1.get_sheet_names():
            if i in self.f2.get_sheet_names() and i not in self.common_sheet:
                self.common_sheet.append(i)

        choose_sheetname = wx.ComboBox(self.panel, choices=self.common_sheet, style=wx.CB_READONLY, pos=(390, 5), size=(100, 25))
        self.Bind(wx.EVT_COMBOBOX, self.cb_sheet, choose_sheetname)

    # -------------------------------------------------------------------
    # def [compareExcel] 比较A、B文件
    def compareExcel(self, event):
        """
        保存单元格行列差异
        :param event:单击compare按钮
        """

        #对比单元格

        self.cmp_cell()

        #对比行/列
        self.cmp_ab_row()
        self.cmp_ab_col()
        self.show_excel()

        #下方改动区域表
        self.gd = wx.grid.Grid(parent=self.panel, id=2, pos=(0, 350), size=(600, 280))
        lg = max(self.dic.keys().__len__(),self.cg_col.keys().__len__(),self.cg_row.keys().__len__())
        self.gd.CreateGrid((lg+1), 3)

    # -------------------------------------------------------------------
    # def [show_excel] 展示A、B文件全部内容
    def show_excel(self):
        """
        展示A、B文件内容并按col输出，不同的地方使用颜色区别
        同步选定状态（A表中的选定单元格在B表同时也被标记）
        """

        # A表按列输出

        self.gd1 = wx.grid.Grid(parent=self.bitmap_1, id=2, pos=(0,70), size=(250, 160))
        col_num_1 = self.f1[self.current_sheet].max_column
        row_num_1 = self.f1[self.current_sheet].max_row

        self.gd1.CreateGrid(row_num_1 + 1,col_num_1,selmode=wx.grid.Grid.SelectCells)

        self.gd1.ClearGrid()

        l1 = self.c_list

        for i in range(len(l1)):
            self.gd1.SetCellValue(0,i,'col-%d'%(i+1))
        ls1 = {}
        for i in range(len(l1)):
            for j in range(1,len(l1[i])+1):
                ls1[(j,i)] = l1[i][j-1]
                if l1[i][j-1] == None:
                    self.gd1.SetCellValue(j, i, '')
                else:
                    self.gd1.SetCellValue(j,i,str(l1[i][j-1]))


        #B表按列输出

        self.gd2 = wx.grid.Grid(parent=self.bitmap_2, id=5, pos=(0,70), size=(250, 160))
        col_num_2 = self.f2[self.current_sheet].max_column
        row_num_2 = self.f2[self.current_sheet].max_row
        self.gd2.CreateGrid(row_num_2 + 1, col_num_2,selmode=wx.grid.Grid.SelectCells)
        self.gd2.ClearGrid()
        l2 = self.d_list

        for i in range(len(l2)):
            self.gd2.SetCellValue(0, i, 'col-%d' % (i + 1))
        ls2 = {}
        for i in range(len(l2)):
            for j in range(1, len(l2[i]) + 1):
                ls2[(j, i)] = l2[i][j - 1]
                if l2[i][j - 1] == None:
                    self.gd2.SetCellValue(j, i, '')
                else:
                    self.gd2.SetCellValue(j, i, str(l2[i][j - 1]))

        #设置单元格颜色
        for key in ls1.keys():
            if ls1[key] != ls2[key]:

                if ls1[key] == None and ls2[key] != None :
                    self.gd1.SetCellBackgroundColour(key[0],key[1],'sky blue')
                    self.gd2.SetCellBackgroundColour(key[0],key[1],'sky blue')
                elif ls1[key] != None  and ls2[key] == None:
                    self.gd1.SetCellBackgroundColour(key[0],key[1],'Red')
                    self.gd2.SetCellBackgroundColour(key[0],key[1],'Red')
                elif ls1[key] != None and ls2 != None and ls1[key] != ls2[key]:
                    self.gd1.SetCellBackgroundColour(key[0], key[1], 'Yellow')
                    self.gd2.SetCellBackgroundColour(key[0], key[1], 'Yellow')

        #选定A表单元格，显示B表单元格位置

        self.gd1.Bind(wx.grid.EVT_GRID_SELECT_CELL, self.onBSelect)

    # -------------------------------------------------------------------
    # def [cmp_cell] 对比A、B文件单元格差异
    def cmp_cell(self):
        """
        同一sheet下对比A、B单元格差异
        """
        ws_a = self.f1.get_sheet_by_name(self.current_sheet)
        ws_b = self.f2.get_sheet_by_name(self.current_sheet)

        r = max(ws_a.max_row,ws_b.max_row)
        c = max(ws_a.max_column,ws_b.max_column)

        for i in range(1, r + 1):
            for j in range(1, c + 1):
                if ws_a.cell(i, j).value != ws_b.cell(i, j).value:
                    # print(ws_a.cell(i, j).value, ws_b.cell(i, j).value)
                    self.dic[(i, j)] = str(ws_a.cell(i, j).value) + '-' + str(ws_b.cell(i, j).value)

    def cb_sheet(self,event):
        current_selection = event.GetSelection()
        self.current_sheet = self.common_sheet[current_selection]
        self.show_bl = True
        self.Refresh()

        #展示文件sheet内容

    # -------------------------------------------------------------------
    # def [cmp_ab_row] 对比A、B文件行增删
    def cmp_ab_row(self):
        """
        同一sheet下对比A、B文件行增删
        删除共有一种情况，新增有两种情况
        """
        a = self.f1.get_sheet_by_name(self.current_sheet)
        b = self.f2.get_sheet_by_name(self.current_sheet)
        #print(a,b)
        # 把表a中的row输入到列表内
        a_list = []
        for row in a.rows:
            tmp_list = []
            a_list.append(tmp_list)
            for cell in row:
                tmp_list.append(cell.value)
        b_list = []
        for row in b.rows:
            tmp_list = []
            b_list.append(tmp_list)
            for cell in row:
                tmp_list.append(cell.value)
        #print(a_list,b_list)
        min_row = (min(len(a_list), len(b_list)))
        max_row = (max(len(a_list), len(b_list)))
        self.cg_row = {}       # 新增/删除行

        for i in range(min_row):
            if operator.eq(a_list[i], b_list[i]) == 1:
                #print(i)
                continue

            # 删除
            if a_list[i].count(None) != len(a_list[i]) and b_list[i].count(None) == len(b_list[i]):
                self.cg_row[str(i+1)] = '删除'
                continue
            # 新增1
            if a_list[i].count(None) == len(a_list[i]) and b_list[i].count(None) != len(b_list[i]):
                self.cg_row[str(i+1)] = '新增'
                continue
        # 新增2
        if max_row != min_row:
            for j in range(min_row + 1, max_row + 1):
                self.cg_row[j] = '新增'

    # -------------------------------------------------------------------
    # def [cmp_ab_col] 对比A、B文件列增删
    def cmp_ab_col(self):
        """
        同一sheet下对比A、B文件行增删
        删除共有一种情况，新增有两种情况
        """
        c = self.f1.get_sheet_by_name(self.current_sheet)
        d = self.f2.get_sheet_by_name(self.current_sheet)

        # 把表a中的col输入到列表内
        self.c_list = []
        for col in c.columns:
            tmp_list = []
            self.c_list.append(tmp_list)
            for cell in col:
                tmp_list.append(cell.value)
        self.d_list = []
        for col in d.columns:
            tmp_list = []
            self.d_list.append(tmp_list)
            for cell in col:
                tmp_list.append(cell.value)
        #print(self.c_list, self.d_list)
        min_col = (min(len(self.c_list), len(self.d_list)))
        max_col = (max(len(self.c_list), len(self.d_list)))

        for i in range(min_col):
            if operator.eq(self.c_list[i], self.d_list[i]) == 1:

                continue

            # 删除
            if self.c_list[i].count(None) != len(self.c_list[i]) and self.d_list[i].count(None) == len(self.d_list[i]):
                self.cg_col[str(i + 1)] = '删除'
                continue
            # 新增1
            if self.c_list[i].count(None) == len(self.c_list[i]) and self.d_list[i].count(None) != len(self.d_list[i]):
                self.cg_col[str(i + 1)] = '新增'
                continue
        # 新增2
        if max_col != min_col:
            for j in range(min_col + 1, max_col + 1):
                self.cg_col[j] = '新增'

    # -------------------------------------------------------------------
    # def [onColSelect] 展示A、B文件行增删
    def onColSelect(self, event):
        """
        选择行增删grid中的内容，相关行会在A、B文件展示grid中被选定标记
        :param event:单击下方差异展示grid中单元格
        """
        # 获取所选行的列值
        v1 = self.gd.GetCellValue(row=event.GetRow(), col=0)

        print(event.GetRow())
        # 如果选中第一行的内容或者表内无差异内容，那么结束
        if event.GetRow() == 0 or self.gd.GetCellValue(event.GetRow(), event.GetCol()) is '':
            pass
        else:

            self.gd1.SelectCol(int(v1)-1,addToSelected = False)
            self.gd2.SelectCol(int(v1)-1,addToSelected = False)

    # -------------------------------------------------------------------
    # def [onRowSelect] 展示A、B文件列增删
    def onRowSelect(self, event):
        """
        选择行增删grid中的内容，相关列会在A、B文件展示grid中被选定标记
        :param event:单击下方差异展示grid中单元格
        """

        # 获取所选列的行值
        v2 = self.gd.GetCellValue(row=event.GetRow(), col=0)

        print(event.GetRow())

        # 如果选中第一行的内容或者表内无差异内容，那么结束
        if event.GetRow() == 0 or self.gd.GetCellValue(event.GetRow(), event.GetCol()) is '':
            pass
        else:
            self.gd1.SelectRow(int(v2) , addToSelected=False)
            self.gd2.SelectRow(int(v2) , addToSelected=False)

    # -------------------------------------------------------------------
    # def [onBSelect] 同步A、B选定单元格
    def onBSelect(self,event):
        """
        同步A、B选定单元格
        :param event:单击下方差异展示grid中单元格
        """

        self.gd2.SelectBlock(topRow = event.GetRow(), leftCol = event.GetCol(), bottomRow = event.GetRow(), rightCol = event.GetCol(), addToSelected=False)

    # -------------------------------------------------------------------
    # def [OnContract] 菜单栏tip
    def OnContract(self,event):
        """
        :param event:单击MainMennu
        """
        wx.MessageBox('If you have any questions, please contact us via E-mail: rocky_chen777@163.com',style=wx.OK,caption='Tips')

    # -------------------------------------------------------------------
    # def [OnContract] 菜单栏tip
    def OnExit(self,event):
        """
        :param event: 单击退出
        :return:
        """
        wx.Exit()

##########################################################################
# Run the program
if __name__ == "__main__":
    app = wx.App(False)
    frame = MyForm()
    frame.Show()

    app.MainLoop()

